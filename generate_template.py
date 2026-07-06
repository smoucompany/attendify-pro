"""
Attendify Pro — Employee Import Template Generator
يولّد قالب Excel احترافي لرفع بيانات الموظفين دفعة واحدة
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import openpyxl.styles.numbers as num_fmt

# ── الألوان ──────────────────────────────────────────────────
C_HEADER_BG    = "1E293B"   # كحلي داكن - خلفية الهيدر
C_HEADER_FG    = "FFFFFF"   # أبيض - نص الهيدر
C_REQUIRED_BG  = "EFF6FF"   # أزرق فاتح جداً - حقول مطلوبة
C_OPTIONAL_BG  = "F8FAFC"   # رمادي فاتح - حقول اختيارية
C_ACCENT       = "6366F1"   # بنفسجي - خط تمييز
C_SAMPLE_BG    = "F0FDF4"   # أخضر فاتح - صف المثال
C_SAMPLE_FG    = "166534"   # أخضر داكن - نص المثال
C_REQUIRED_HDR = "3B82F6"   # أزرق - مؤشر مطلوب
C_OPTIONAL_HDR = "94A3B8"   # رمادي - مؤشر اختياري
C_STRIPE_1     = "FFFFFF"   # أبيض
C_STRIPE_2     = "F8FAFC"   # رمادي خفيف
C_BORDER       = "CBD5E1"   # رمادي - حدود
C_TITLE_BG     = "0F172A"   # أسود مائل - خلفية العنوان
C_SUBHDR_BG    = "334155"   # رمادي داكن - هيدر فرعي
C_TAG_REQ      = "DBEAFE"   # أزرق خفيف - badge مطلوب
C_TAG_OPT      = "F1F5F9"   # رمادي خفيف - badge اختياري

def hex_fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    m = Side(style="medium", color=C_ACCENT)
    return Border(left=m, right=m, top=m, bottom=m)

def make_font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Segoe UI", bold=bold, size=size,
                color=color, italic=italic)

wb = Workbook()

# ════════════════════════════════════════════════════════════════
#  الشيت الأول: قالب الإدخال
# ════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📋 بيانات الموظفين"
ws.sheet_view.rightToLeft = True
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A6"  # تثبيت الهيدر

# ── الصف 1: شريط العنوان ─────────────────────────────────────
ws.merge_cells("A1:M1")
ws.row_dimensions[1].height = 48
title_cell = ws["A1"]
title_cell.value = "🏢  Attendify Pro  —  قالب رفع بيانات الموظفين"
title_cell.font   = Font(name="Segoe UI", bold=True, size=16,
                         color="FFFFFF")
title_cell.fill   = hex_fill(C_TITLE_BG)
title_cell.alignment = Alignment(horizontal="center",
                                  vertical="center",
                                  readingOrder=2)

# ── الصف 2: وصف فرعي ─────────────────────────────────────────
ws.merge_cells("A2:M2")
ws.row_dimensions[2].height = 22
sub = ws["A2"]
sub.value = "أضف بيانات الموظفين ابتداءً من الصف 6  •  لا تحذف الأعمدة  •  الحقول باللون الأزرق مطلوبة"
sub.font  = Font(name="Segoe UI", size=9, color="94A3B8", italic=True)
sub.fill  = hex_fill(C_SUBHDR_BG)
sub.alignment = Alignment(horizontal="center", vertical="center",
                           readingOrder=2)

# ── الصف 3: فارغ مع خط ────────────────────────────────────────
ws.row_dimensions[3].height = 4
for col in range(1, 14):
    ws.cell(row=3, column=col).fill = hex_fill(C_ACCENT)

# ── الصف 4: أسطورة (legend) ──────────────────────────────────
ws.row_dimensions[4].height = 20
ws.merge_cells("A4:C4")
leg1 = ws["A4"]
leg1.value = "🔵  مطلوب — يجب تعبئته"
leg1.font  = Font(name="Segoe UI", size=8.5, color="1D4ED8")
leg1.fill  = hex_fill("DBEAFE")
leg1.alignment = Alignment(horizontal="center", vertical="center",
                            readingOrder=2)

ws.merge_cells("D4:F4")
leg2 = ws["D4"]
leg2.value = "⚪  اختياري — يمكن تركه فارغاً"
leg2.font  = Font(name="Segoe UI", size=8.5, color="475569")
leg2.fill  = hex_fill("F1F5F9")
leg2.alignment = Alignment(horizontal="center", vertical="center",
                            readingOrder=2)

ws.merge_cells("G4:I4")
leg3 = ws["G4"]
leg3.value = "🟢  الصف الأخضر — مثال توضيحي (احذفه عند الرفع)"
leg3.font  = Font(name="Segoe UI", size=8.5, color="166534")
leg3.fill  = hex_fill("DCFCE7")
leg3.alignment = Alignment(horizontal="center", vertical="center",
                            readingOrder=2)

ws.merge_cells("J4:M4")
leg4 = ws["J4"]
leg4.value = "🔽  الخلايا تحتوي قوائم منسدلة للاختيار"
leg4.font  = Font(name="Segoe UI", size=8.5, color="7C3AED")
leg4.fill  = hex_fill("EDE9FE")
leg4.alignment = Alignment(horizontal="center", vertical="center",
                            readingOrder=2)

# ── الصف 5: فارغ ──────────────────────────────────────────────
ws.row_dimensions[5].height = 6

# ── تعريف الأعمدة ──────────────────────────────────────────────
COLUMNS = [
    # (header, width, required, hint)
    ("الاسم الأول",        18, True,  "أحمد"),
    ("اسم العائلة",        18, True,  "محمد"),
    ("رقم الهاتف",         16, False, "0501234567"),
    ("البريد الإلكتروني",  24, False, "ahmed@company.com"),
    ("القسم",              18, True,  "الموارد البشرية"),
    ("المنصب / الوظيفة",   20, False, "محاسب"),
    ("الراتب الأساسي",     16, True,  "5000"),
    ("تاريخ التعيين",      16, True,  "2024-01-15"),
    ("الجنس",              12, False, "ذكر"),
    ("الحالة",             12, False, "نشط"),
    ("رقم الموظف",         14, False, "EMP-001"),
    ("ملاحظات",            24, False, ""),
    ("رمز الدخول (اختياري)", 18, False, ""),
]

# ── الصف 6: هيدر الأعمدة ─────────────────────────────────────
ws.row_dimensions[6].height = 38

for col_idx, (header, width, required, _) in enumerate(COLUMNS, start=1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

    cell = ws.cell(row=6, column=col_idx)
    cell.value = header
    cell.font  = Font(name="Segoe UI", bold=True, size=10,
                      color="FFFFFF")
    cell.fill  = hex_fill(C_HEADER_BG if required else C_SUBHDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, readingOrder=2)
    cell.border = thin_border("475569")

    # خط ملون أسفل الهيدر
    ws.cell(row=7, column=col_idx).fill = hex_fill(
        C_REQUIRED_HDR if required else C_OPTIONAL_HDR
    )

ws.row_dimensions[7].height = 3

# ── الصف 7 (=8 للبيانات): صف المثال ─────────────────────────
SAMPLE = [
    "أحمد", "محمد", "0501234567", "ahmed@company.com",
    "الموارد البشرية", "محاسب", 5000, "2024-01-15",
    "ذكر", "نشط", "EMP-001", "موظف نموذجي", ""
]

ws.row_dimensions[8].height = 28
for col_idx, value in enumerate(SAMPLE, start=1):
    cell = ws.cell(row=8, column=col_idx)
    cell.value = value
    cell.font  = Font(name="Segoe UI", size=9.5,
                      color=C_SAMPLE_FG, italic=True)
    cell.fill  = hex_fill(C_SAMPLE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                readingOrder=2)
    cell.border = thin_border("86EFAC")

# ── الصفوف 9–108: صفوف البيانات ──────────────────────────────
for row in range(9, 109):
    stripe = C_STRIPE_1 if row % 2 == 0 else C_STRIPE_2
    ws.row_dimensions[row].height = 26
    for col_idx, (_, _, required, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = hex_fill(
            C_REQUIRED_BG if required else stripe
        )
        cell.alignment = Alignment(horizontal="center",
                                    vertical="center",
                                    readingOrder=2)
        cell.border = thin_border()
        cell.font   = Font(name="Segoe UI", size=10)

# ── Validation: الجنس ─────────────────────────────────────────
dv_gender = DataValidation(
    type="list",
    formula1='"ذكر,أنثى"',
    allow_blank=True,
    showDropDown=False,
    error='اختر من القائمة: ذكر أو أنثى',
    errorTitle='قيمة غير صالحة',
    prompt='اختر الجنس',
    promptTitle='الجنس'
)
ws.add_data_validation(dv_gender)
dv_gender.sqref = "I9:I108"

# ── Validation: الحالة ────────────────────────────────────────
dv_status = DataValidation(
    type="list",
    formula1='"نشط,غير نشط,في إجازة"',
    allow_blank=True,
    showDropDown=False,
    error='اختر: نشط / غير نشط / في إجازة',
    errorTitle='قيمة غير صالحة',
    prompt='اختر الحالة الوظيفية',
    promptTitle='الحالة'
)
ws.add_data_validation(dv_status)
dv_status.sqref = "J9:J108"

# ── Validation: التاريخ ───────────────────────────────────────
dv_date = DataValidation(
    type="date",
    operator="greaterThan",
    formula1="DATE(2000,1,1)",
    allow_blank=True,
    error='أدخل تاريخاً صحيحاً (مثال: 2024-01-15)',
    errorTitle='تاريخ غير صالح',
    prompt='أدخل بالصيغة: YYYY-MM-DD',
    promptTitle='تاريخ التعيين'
)
ws.add_data_validation(dv_date)
dv_date.sqref = "H9:H108"

# ── Validation: الراتب (رقم موجب) ────────────────────────────
dv_salary = DataValidation(
    type="decimal",
    operator="greaterThan",
    formula1="0",
    allow_blank=True,
    error='أدخل رقماً أكبر من صفر',
    errorTitle='راتب غير صالح',
    prompt='أدخل الراتب بالأرقام فقط',
    promptTitle='الراتب الأساسي'
)
ws.add_data_validation(dv_salary)
dv_salary.sqref = "G9:G108"

# ── رقم الهاتف (نص) ─────────────────────────────────────────
for row in range(9, 109):
    cell = ws.cell(row=row, column=3)
    cell.number_format = "@"  # نص

# ── الصف الأخير: تذييل ───────────────────────────────────────
last_row = 109
ws.merge_cells(f"A{last_row}:M{last_row}")
footer = ws.cell(row=last_row, column=1)
footer.value = (
    "© Attendify Pro  |  يدعم 100 موظف في هذا القالب  "
    "—  للمزيد انسخ الصفوف  |  احذف صف المثال الأخضر قبل الرفع"
)
footer.font  = Font(name="Segoe UI", size=8, color="94A3B8", italic=True)
footer.fill  = hex_fill(C_TITLE_BG)
footer.alignment = Alignment(horizontal="center", vertical="center",
                              readingOrder=2)
ws.row_dimensions[last_row].height = 22


# ════════════════════════════════════════════════════════════════
#  الشيت الثاني: تعليمات الرفع
# ════════════════════════════════════════════════════════════════
wi = wb.create_sheet("📖 تعليمات الرفع")
wi.sheet_view.rightToLeft = True
wi.sheet_view.showGridLines = False
wi.column_dimensions["A"].width = 4
wi.column_dimensions["B"].width = 32
wi.column_dimensions["C"].width = 55

# عنوان
wi.merge_cells("A1:C1")
wi.row_dimensions[1].height = 44
t = wi["A1"]
t.value = "📖  دليل رفع بيانات الموظفين"
t.font  = Font(name="Segoe UI", bold=True, size=15, color="FFFFFF")
t.fill  = hex_fill(C_TITLE_BG)
t.alignment = Alignment(horizontal="center", vertical="center",
                         readingOrder=2)

wi.row_dimensions[2].height = 4
for c in range(1,4):
    wi.cell(row=2, column=c).fill = hex_fill(C_ACCENT)

# دالة لإضافة سطر تعليمات
row_i = [3]
def add_section(title, color_bg="1E293B", color_fg="FFFFFF"):
    wi.row_dimensions[row_i[0]].height = 6
    row_i[0] += 1
    wi.merge_cells(f"A{row_i[0]}:C{row_i[0]}")
    wi.row_dimensions[row_i[0]].height = 30
    c = wi.cell(row=row_i[0], column=1)
    c.value = f"  {title}"
    c.font  = Font(name="Segoe UI", bold=True, size=11, color=color_fg)
    c.fill  = hex_fill(color_bg)
    c.alignment = Alignment(vertical="center", readingOrder=2)
    row_i[0] += 1

def add_row(icon, field, desc, required=False):
    wi.row_dimensions[row_i[0]].height = 24
    ic = wi.cell(row=row_i[0], column=1)
    ic.value = icon
    ic.alignment = Alignment(horizontal="center", vertical="center")
    ic.fill = hex_fill(C_STRIPE_2)

    fc = wi.cell(row=row_i[0], column=2)
    fc.value = field
    fc.font  = Font(name="Segoe UI", bold=required, size=10,
                    color="1D4ED8" if required else "334155")
    fc.fill  = hex_fill(C_REQUIRED_BG if required else C_STRIPE_2)
    fc.alignment = Alignment(vertical="center", readingOrder=2,
                              indent=1)
    fc.border = thin_border()

    dc = wi.cell(row=row_i[0], column=3)
    dc.value = desc
    dc.font  = Font(name="Segoe UI", size=9.5, color="475569")
    dc.fill  = hex_fill(C_STRIPE_1 if required else C_STRIPE_2)
    dc.alignment = Alignment(vertical="center", wrap_text=True,
                              readingOrder=2, indent=1)
    dc.border = thin_border()
    row_i[0] += 1

add_section("📋  الحقول والتعليمات", "1E293B", "FFFFFF")
add_row("🔵", "الاسم الأول *",          "مطلوب — اسم الموظف الأول", True)
add_row("🔵", "اسم العائلة *",          "مطلوب — اسم العائلة", True)
add_row("🔵", "القسم *",               "مطلوب — يجب أن يتطابق مع أقسام النظام", True)
add_row("🔵", "الراتب الأساسي *",      "مطلوب — أرقام فقط بدون رموز عملة", True)
add_row("🔵", "تاريخ التعيين *",        "مطلوب — بالصيغة YYYY-MM-DD (مثال: 2024-03-15)", True)
add_row("⚪", "رقم الهاتف",             "اختياري — يفضل مع رمز الدولة (966...)")
add_row("⚪", "البريد الإلكتروني",      "اختياري — للتواصل والإشعارات")
add_row("⚪", "المنصب / الوظيفة",       "اختياري — مسمى الوظيفة")
add_row("🔽", "الجنس",                 "اختر من القائمة: ذكر / أنثى")
add_row("🔽", "الحالة",                "اختر: نشط / غير نشط / في إجازة (الافتراضي: نشط)")
add_row("⚪", "رقم الموظف",             "اختياري — سيُولَّد تلقائياً إن تُرك فارغاً")
add_row("⚪", "ملاحظات",               "اختياري — أي ملاحظات إضافية")
add_row("⚪", "رمز الدخول",             "اختياري — كلمة مرور المنظومة الداخلية (بوابة الموظف)")

add_section("⚡  خطوات الرفع", "334155", "FFFFFF")
steps = [
    ("1️⃣", "تعبئة البيانات",     "أدخل بيانات موظفيك ابتداءً من الصف 9"),
    ("2️⃣", "حذف صف المثال",     "احذف الصف الأخضر (الصف 8) قبل الرفع"),
    ("3️⃣", "حفظ الملف",          "احفظ الملف بامتداد .xlsx"),
    ("4️⃣", "فتح الموقع",         "اذهب إلى قسم الموظفين في النظام"),
    ("5️⃣", "رفع الملف",          "اضغط 'رفع جماعي' ← اختر الملف ← تأكيد"),
]
for icon, field, desc in steps:
    add_row(icon, field, desc)

add_section("⚠️  تحذيرات مهمة", "7F1D1D", "FCA5A5")
warnings = [
    ("❌", "لا تغير أسماء الأعمدة",     "أي تعديل في الهيدر يوقف عملية الرفع"),
    ("❌", "لا تدمج الخلايا",            "الخلايا المدموجة تسبب أخطاء في الاستيراد"),
    ("❌", "لا تترك الصف 6 فارغاً",     "صف الهيدر ضروري لتحديد الأعمدة"),
    ("✅", "يمكن ترك الحقول الاختيارية", "فارغة — لن يتسبب ذلك في أخطاء"),
    ("✅", "الملف يدعم حتى 100 موظف",   "لبيانات أكثر: انسخ الصفوف لأسفل"),
]
for icon, field, desc in warnings:
    add_row(icon, field, desc)

# تذييل الشيت الثاني
row_i[0] += 1
wi.merge_cells(f"A{row_i[0]}:C{row_i[0]}")
wi.row_dimensions[row_i[0]].height = 22
f2 = wi.cell(row=row_i[0], column=1)
f2.value = "Attendify Pro  —  نظام الحضور والانصراف الذكي"
f2.font  = Font(name="Segoe UI", size=8, color="94A3B8", italic=True)
f2.fill  = hex_fill(C_TITLE_BG)
f2.alignment = Alignment(horizontal="center", vertical="center",
                          readingOrder=2)


# ════════════════════════════════════════════════════════════════
#  الشيت الثالث: الأقسام المرجعية
# ════════════════════════════════════════════════════════════════
wr = wb.create_sheet("📊 قوائم مرجعية")
wr.sheet_view.rightToLeft = True
wr.sheet_view.showGridLines = False

wr.merge_cells("A1:D1")
wr.row_dimensions[1].height = 36
ref_title = wr["A1"]
ref_title.value = "📊  القوائم المرجعية — للاستيراد الصحيح"
ref_title.font  = Font(name="Segoe UI", bold=True, size=13, color="FFFFFF")
ref_title.fill  = hex_fill(C_TITLE_BG)
ref_title.alignment = Alignment(horizontal="center", vertical="center",
                                 readingOrder=2)

wr.row_dimensions[2].height = 4
for c in range(1,5):
    wr.cell(row=2, column=c).fill = hex_fill(C_ACCENT)

# أعمدة
for col, (hdr, w) in enumerate([
    ("الجنس", 16), ("الحالة الوظيفية", 22),
    ("أمثلة الأقسام", 26), ("صيغة التاريخ", 22)
], start=1):
    wr.column_dimensions[get_column_letter(col)].width = w
    c = wr.cell(row=3, column=col)
    c.value = hdr
    c.font  = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    c.fill  = hex_fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             readingOrder=2)
    c.border = thin_border("475569")
wr.row_dimensions[3].height = 30

ref_data = [
    ("ذكر",    "نشط",         "الموارد البشرية",    "2024-01-15 ✅"),
    ("أنثى",   "غير نشط",     "المحاسبة والمالية",  "2024/01/15 ❌"),
    ("",       "في إجازة",    "المبيعات والتسويق",  "15-01-2024 ❌"),
    ("",       "",            "تقنية المعلومات",    "15/1/2024  ❌"),
    ("",       "",            "العمليات والإنتاج",  ""),
    ("",       "",            "خدمة العملاء",       ""),
    ("",       "",            "الإدارة العليا",     ""),
]

fills_ref = [C_STRIPE_1, C_STRIPE_2]
for i, row_data in enumerate(ref_data, start=4):
    wr.row_dimensions[i].height = 24
    for col, val in enumerate(row_data, start=1):
        c = wr.cell(row=i, column=col)
        c.value = val
        c.font  = Font(name="Segoe UI", size=10)
        c.fill  = hex_fill(fills_ref[i % 2])
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 readingOrder=2)
        c.border = thin_border()

# ── حفظ الملف ─────────────────────────────────────────────────
output_path = r"d:\employee_template.xlsx"
wb.save(output_path)
import sys
sys.stdout.buffer.write(b"Done: d:\\employee_template.xlsx\n")
